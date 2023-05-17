from pyproj import Proj, transform
import datetime
import pyproj
import openpyxl
import warnings
import sys
import os
import time
import string
import tkinter as tk
from tkinter import filedialog

#terminal = subprocess.Popen('start cmd', stdout=subprocess.PIPE, stdin=subprocess.PIPE, shell=True)
#sys.stdout = terminal.stdin

start_time = time.time() #Inicio timer
link = "https://github.com/EdwinForero/ProyectoDatosRioBogota"

salida = "\n------------- CONVERSOR DE COORDENADAS A WSG84 EPSG:4326 ------------" \
    + "\nEPSG:3116 MAGNA-SIRGAS / Colombia Bogota zone" \
    + "\nESRI:103599 MAGNA-SIRGAS CMT12" \
    + "\n\n--Developed by: Edwin Forero & David Forero" \
    + "\nPara mayor info, consulte: " \
    + link \
    + "\n______________________________________________________________________\n"

print(salida)

warnings.filterwarnings("ignore") #Para no imprimir warnings

relaciones = {} #Diccionario columnas excel
letras = string.ascii_uppercase  # Generar todas las letras mayúsculas de la A a la Z

# Generar todas las combinaciones de dos letras de AA a ZZ
for letra1 in letras:
    for letra2 in letras:
        llave = letra1 + letra2
        valor = (string.ascii_uppercase.index(letra1) + 1) * 26 + string.ascii_uppercase.index(letra2) + 1
        relaciones[valor] = llave

# Agregar las letras de la A a la Z con sus valores correspondientes
for i, letra in enumerate(letras):
    relaciones[i+1] = letra

#Rutas
pathA = hojaA = ""

#Llamamos al explorador
root = tk.Tk()
root.withdraw()

# Llamamos al explorador de archivos y guardamos la ruta del archivo seleccionado
pathA = filedialog.askopenfilename(filetypes=[('Excel files', '*.xlsx;*.xlsb;*.xlsm'), ('CSV files', '*.csv')])

#---Si no se seleccionó un archivo salimos
if pathA == None or pathA == "":
    print("\nNo ha seleccionado un archivo, saliendo del programa...")
    time.sleep(5)
    sys.exit(1)

#---Flujo normal
print("\n---Leyendo Archivo:", pathA, "\nEspere un momento...")

# Abrir el archivo de Excel
workbook = openpyxl.load_workbook(pathA)
hojaA = workbook.sheetnames[0]
# Seleccionar la hoja de trabajo
sheet = workbook[hojaA]

#Estado
print("\nRuta:", pathA)
print("\nHoja:", hojaA)
print("\n------Archivo OK------")

# Definir los sistemas de referencia de coordenadas de origen y destino
#EPSG:3116 MAGNA-SIRGAS / Colombia Bogota zone
origen = pyproj.Proj('EPSG:3116') 
destino = pyproj.Proj('EPSG:4326') #WSG84

#ESRI:103599 MAGNA-SIRGAS CMT12 #2M
esri_crs = pyproj.CRS('ESRI:103599')
wgs84_crs = pyproj.CRS('WGS84')


#Preguntar columna origen X,Y
numX = input("\nIngrese número de columna para leer las coordenadas en X (1-52): ")
while not numX.isdigit() or int(numX) < 1 or int(numX) > 52:
    print("\n\nEjemplo: A-1; N-14; P-16; Y-25; AA-27; AZ-52")
    numX = input("\nIngrese número de columna para leer las coordenadas en X (1-52): ")

numY = input("\nIngrese número de columna para leer las coordenadas en Y (1-52): ")
while not numY.isdigit() or int(numY) < 1 or int(numY) > 52:
    print("\n\nEjemplo: A-1; N-14; P-16; Y-25; AA-27; AZ-52")
    numY = input("\nIngrese número de columna para leer las coordenadas en Y (1-52): ")

# Leer y guardas los valores de las columnas 
column_o = sheet[relaciones[int(numY)]]
column_p = sheet[relaciones[int(numX)]]

print("\n\nColumnas origen guardadas.")

#Preguntar columna destino X,Y
desX = input("\nIngrese número de columna donde se guardará la Latitud (1-52): ")
while not desX.isdigit() or int(desX) < 1 or int(desX) > 52:
    print("\n\nEjemplo: A-1; N-14; P-16; Y-25; AA-27; AZ-52")
    desX = input("\nIngrese número de columna donde se guardará la Latitud (1-52): ")

desY = input("\nIngrese número de columna donde se guardará la Longitud (1-52): ")
while not desY.isdigit() or int(desY) < 1 or int(desY) > 52:
    print("\n\nEjemplo: A-1; N-14; P-16; Y-25; AA-27; AZ-52")
    desY = input("\nIngrese número de columna donde se guardará la Longitud (1-52): ")

sheet.cell(row=1, column=int(desY)).value = "Longitud" # Encabezado fila 1 Y
sheet.cell(row=1, column=int(desX)).value = "Latitud"  # Encabezado fila 1 X

print("\n\nColumnas destino capturadas.")

# Guardar los cambios en el archivo de Excel
print("\nVerificando...")
workbook.save(pathA)

print("\n------Inicio OK------")
print("\nConvirtiendo coordenadas...")

errorsN = []

#Converir coordenadas
for i in range(2, len(column_o)+1):
    x = y = lon = lat = x_m = y_m = 0
    
    try:
        x = float(column_o[i-1].value) #y
        y = float(column_p[i-1].value) #x
    except: #Si no se puede parsear
        errorsN.append(i)
        continue #Saltamos linea
    
    #Si la coordenada es 0 o -1 la omitimos
    if x == float(0) or x == float(-1):
        errorsN.append(i)
        continue

    # Casos
    if x >= 2000000: #ESRI:103599 MAGNA-SIRGAS CMT12 
        transformer = pyproj.Transformer.from_crs(esri_crs, wgs84_crs)
        x_m, y_m = transformer.transform(y, x)
        lat = x_m
        lon = y_m
    else: #EPSG:3116 MAGNA-SIRGAS / Colombia Bogota zone
        x_m, y_m = pyproj.transform(origen, destino, x, y)
        lon, lat = pyproj.transform(pyproj.Proj(proj='latlong', datum='WGS84'), destino, x_m, y_m)


    sheet.cell(row=i, column=int(desX)).value = lat  # Guardamos latitud en columna
    sheet.cell(row=i, column=int(desY)).value = lon # Guardamos longitud en columna


    print("\rFila ",i, " de ", len(column_o), end="", flush=True) #Estado
    sys.stdout.flush()

  
# Imprimimos filas saltadas
print("\n\nTotal filas saltadas", len(errorsN))  

# Guardar los cambios en el archivo de Excel
print("\nGuardando cambios...")  
workbook.save(pathA)

desY = input("\n¿Desea ver cuáles filas se omitieron? Digite '1': ")
if desY.isdigit() and int(desY) == 1:
    print("\n--------Filas saltadas--------\n")
    print(errorsN)

# Fin timer
end_time = time.time()

total_time = end_time - start_time

horas=int(total_time/3600)

total_time-=horas*3600

minutos=int(total_time/60)

total_time-=minutos*60

# Impresiones finales
print("\n\nTiempo total de ejecución: {:02d}:{:02d}:{:02d}".format(horas, minutos, int(total_time)))
print("\n------Fin del proceso------")

print("\n\nPresione cualquier tecla para salir...")
input()  # espera a que el usuario presione una tecla
